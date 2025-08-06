from fastapi import APIRouter, File, UploadFile, HTTPException, Depends
from fastapi.responses import JSONResponse
import pandas as pd
import openpyxl
import io
import logging
from typing import List, Dict, Any
from pydantic import BaseModel

from auth import get_current_user
from models import APIResponse, User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["excel-compare"])


class ExcelCompareResult(BaseModel):
    """Result model for Excel comparison."""
    success: bool
    total_sheets: int
    matched_sheets: int
    comparison_results: List[Dict[str, Any]]
    summary: str


@router.post("/excel-compare", response_model=APIResponse)
async def compare_excel_files(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Compare two Excel files sheet by sheet, cell by cell.
    
    All users can access this functionality.
    Both files must have the same number of sheets with matching sheet names.
    """
    
    try:
        # Validate file types
        if not (file1.filename.lower().endswith(('.xlsx', '.xls')) and 
                file2.filename.lower().endswith(('.xlsx', '.xls'))):
            raise HTTPException(status_code=400, detail="Both files must be Excel files (.xlsx or .xls)")
        
        # Read Excel files
        file1_content = await file1.read()
        file2_content = await file2.read()
        
        workbook1 = openpyxl.load_workbook(io.BytesIO(file1_content), data_only=True)
        workbook2 = openpyxl.load_workbook(io.BytesIO(file2_content), data_only=True)
        
        # Get sheet names
        sheets1 = workbook1.sheetnames
        sheets2 = workbook2.sheetnames
        
        # Handle different number of sheets gracefully
        if len(sheets1) != len(sheets2):
            logger.warning(f"Files have different number of sheets: {len(sheets1)} vs {len(sheets2)}")
            # Still proceed with comparison of common sheets
            common_sheets = set(sheets1) & set(sheets2)
            if not common_sheets:
                raise HTTPException(
                    status_code=400,
                    detail=f"No common sheets found between files. File 1 has: {sheets1}. File 2 has: {sheets2}"
                )
            # Use common sheets for comparison
            sheets_to_compare = list(common_sheets)
            logger.info(f"Comparing {len(sheets_to_compare)} common sheets: {sheets_to_compare}")
        else:
            sheets_to_compare = sheets1
        
        # Note: We now handle mismatched sheets gracefully above
        
        comparison_results = []
        matched_sheets = 0
        
        # Compare each sheet
        for sheet_name in sheets_to_compare:
            if sheet_name in workbook1.sheetnames and sheet_name in workbook2.sheetnames:
                sheet_result = compare_sheets(
                    workbook1[sheet_name], 
                    workbook2[sheet_name], 
                    sheet_name
                )
                comparison_results.append(sheet_result)
                
                if sheet_result["status"] == "matched":
                    matched_sheets += 1
            else:
                # Sheet exists in one file but not the other
                comparison_results.append({
                    "sheet": sheet_name,
                    "cell_id": "NA",
                    "value1": "sheet exists" if sheet_name in workbook1.sheetnames else "sheet missing",
                    "value2": "sheet exists" if sheet_name in workbook2.sheetnames else "sheet missing",
                    "status": "sheet_mismatch",
                    "differences": []
                })
        
        # Add info about sheets that exist in only one file
        sheets1_only = set(sheets1) - set(sheets2)
        sheets2_only = set(sheets2) - set(sheets1)
        
        for sheet_name in sheets1_only:
            comparison_results.append({
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "sheet exists",
                "value2": "sheet missing",
                "status": "sheet_missing_in_file2",
                "differences": []
            })
            
        for sheet_name in sheets2_only:
            comparison_results.append({
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "sheet missing",
                "value2": "sheet exists", 
                "status": "sheet_missing_in_file1",
                "differences": []
            })
        
        success = matched_sheets == len(sheets_to_compare) and len(sheets1_only) == 0 and len(sheets2_only) == 0
        total_sheets_compared = len(sheets_to_compare) + len(sheets1_only) + len(sheets2_only)
        summary = f"Compared {total_sheets_compared} sheets. {matched_sheets} matched, {total_sheets_compared - matched_sheets} had differences or were missing."
        
        result = {
            "success": success,
            "total_sheets": total_sheets_compared,
            "matched_sheets": matched_sheets,
            "comparison_results": comparison_results,
            "summary": summary,
            "files_info": {
                "file1_name": file1.filename,
                "file2_name": file2.filename,
                "file1_sheets": len(sheets1),
                "file2_sheets": len(sheets2)
            }
        }
        
        return APIResponse(
            success=True,
            message="Excel comparison completed successfully",
            data=result
        )
        
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Excel comparison error: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to compare Excel files: {str(exc)}")


def compare_sheets(sheet1, sheet2, sheet_name: str) -> Dict[str, Any]:
    """Compare two Excel sheets cell by cell."""
    
    differences = []
    
    # Get the maximum row and column from both sheets
    max_row1, max_col1 = sheet1.max_row, sheet1.max_column
    max_row2, max_col2 = sheet2.max_row, sheet2.max_column
    
    max_row = max(max_row1, max_row2)
    max_col = max(max_col1, max_col2)
    
    # Compare each cell
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell1_value = None
            cell2_value = None
            
            if row <= max_row1 and col <= max_col1:
                cell1 = sheet1.cell(row=row, column=col)
                cell1_value = cell1.value
            
            if row <= max_row2 and col <= max_col2:
                cell2 = sheet2.cell(row=row, column=col)
                cell2_value = cell2.value
            
            # Normalize None values to empty string for comparison
            if cell1_value is None:
                cell1_value = ""
            if cell2_value is None:
                cell2_value = ""
            
            # Convert to string for comparison
            cell1_str = str(cell1_value) if cell1_value != "" else ""
            cell2_str = str(cell2_value) if cell2_value != "" else ""
            
            if cell1_str != cell2_str:
                cell_id = f"{openpyxl.utils.get_column_letter(col)}{row}"
                differences.append({
                    "sheet": sheet_name,
                    "cell_id": cell_id,
                    "value1": cell1_str,
                    "value2": cell2_str,
                    "status": "not matched"
                })
    
    if not differences:
        return {
            "sheet": sheet_name,
            "cell_id": "NA",
            "value1": "NA", 
            "value2": "NA",
            "status": "matched",
            "differences": []
        }
    else:
        return {
            "sheet": sheet_name,
            "cell_id": "multiple",
            "value1": "multiple",
            "value2": "multiple", 
            "status": "not matched",
            "differences": differences
        }